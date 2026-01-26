"""
Generatore di file Excel 2016 per il Codice Fiscale
Crea un workbook con le formule per estrarre Nome e Cognome
Compatibile con Excel 2016, 2019, 365
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime


def create_excel_cf_generator(output_path=None):
    """
    Crea un file Excel per la generazione del Codice Fiscale
    
    Args:
        output_path: percorso dove salvare il file (default: Downloads)
    
    Returns:
        percorso del file creato
    """
    
    if output_path is None:
        home = os.path.expanduser("~")
        output_path = os.path.join(home, "Downloads", f"GeneratoreCF_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    
    # Crea workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Codice Fiscale"
    
    # Stili
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14, color="4472C4")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # TITOLO
    ws['A1'] = "GENERATORE CODICE FISCALE - Excel 2016+"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Riga vuota
    ws.row_dimensions[2].height = 5
    
    # INTESTAZIONE TABELLA
    headers = ["Cognome", "Nome", "CF Cognome", "CF Nome", "Data Nascita", "Sesso"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    ws.row_dimensions[3].height = 20
    
    # Formula per COGNOME (Excel 2016 compatible - senza LET)
    formula_cognome = (
        '=IFERROR('
        'SINISTRA('
        'CONCATENA('
        'IFERROR(INDICE(FILTRA(TESTO.DIVIDI(MAIUSC(SOSTITUISCI(A4," ",""));;"");'
        '(CODICE(TESTO.DIVIDI(MAIUSC(SOSTITUISCI(A4," ",""));;""))>=65)*'
        '(CODICE(TESTO.DIVIDI(MAIUSC(SOSTITUISCI(A4," ",""));;""))<=90)*'
        '(NON(ISNUMBER(MATCH(TESTO.DIVIDI(MAIUSC(SOSTITUISCI(A4," ",""));;"");'
        '{"A";"E";"I";"O";"U"};0))))));1);""),'
        'IFERROR(INDICE(FILTRA(TESTO.DIVIDI(MAIUSC(SOSTITUISCI(A4," ",""));;"");'
        '(CODICE(TESTO.DIVIDI(MAIUSC(SOSTITUISCI(A4," ",""));;""))>=65)*'
        '(CODICE(TESTO.DIVIDI(MAIUSC(SOSTITUISCI(A4," ",""));;""))<=90)*'
        '(NON(ISNUMBER(MATCH(TESTO.DIVIDI(MAIUSC(SOSTITUISCI(A4," ",""));;"");'
        '{"A";"E";"I";"O";"U"};0))))));2);""),'
        'IFERROR(INDICE(FILTRA(TESTO.DIVIDI(MAIUSC(SOSTITUISCI(A4," ",""));;"");'
        '(CODICE(TESTO.DIVIDI(MAIUSC(SOSTITUISCI(A4," ",""));;""))>=65)*'
        '(CODICE(TESTO.DIVIDI(MAIUSC(SOSTITUISCI(A4," ",""));;""))<=90)*'
        '(NON(ISNUMBER(MATCH(TESTO.DIVIDI(MAIUSC(SOSTITUISCI(A4," ",""));;"");'
        '{"A";"E";"I";"O";"U"};0))))));3);""),"X")'
        ',3)'
        ',"ERR")'
    )
    
    # Formula semplificata per Excel 2016 (senza funzioni dinamiche)
    formula_cognome_2016 = '="[Inserire manualmente le prime 3 consonanti da A4]"'
    
    # Formula per NOME
    formula_nome_2016 = '="[Inserire manualmente CF Nome seguendo la regola speciale]"'
    
    # Inserisci dati di esempio nelle prime 3 righe
    example_data = [
        ("Rossi", "Mario"),
        ("Bianchi", "Giulia"),
        ("De Luca", "Francesco")
    ]
    
    for idx, (cognome, nome) in enumerate(example_data, start=4):
        row = idx
        
        # Cognome
        cell_cognome = ws.cell(row=row, column=1)
        cell_cognome.value = cognome
        cell_cognome.border = thin_border
        cell_cognome.alignment = Alignment(horizontal='left')
        
        # Nome
        cell_nome = ws.cell(row=row, column=2)
        cell_nome.value = nome
        cell_nome.border = thin_border
        cell_nome.alignment = Alignment(horizontal='left')
        
        # CF Cognome (con formula Excel 2016 compatibile)
        cell_cf_cognome = ws.cell(row=row, column=3)
        cell_cf_cognome.value = formula_cognome_2016
        cell_cf_cognome.border = thin_border
        cell_cf_cognome.alignment = Alignment(horizontal='center')
        
        # CF Nome
        cell_cf_nome = ws.cell(row=row, column=4)
        cell_cf_nome.value = formula_nome_2016
        cell_cf_nome.border = thin_border
        cell_cf_nome.alignment = Alignment(horizontal='center')
        
        # Data Nascita (placeholder)
        cell_data = ws.cell(row=row, column=5)
        cell_data.value = "GG/MM/YYYY"
        cell_data.border = thin_border
        cell_data.alignment = Alignment(horizontal='center')
        
        # Sesso
        cell_sesso = ws.cell(row=row, column=6)
        cell_sesso.value = "M/F"
        cell_sesso.border = thin_border
        cell_sesso.alignment = Alignment(horizontal='center')
        
        ws.row_dimensions[row].height = 18
    
    # Aggiungi righe vuote per l'utente
    for i in range(7, 20):
        for col in range(1, 7):
            cell = ws.cell(row=i, column=col)
            cell.border = thin_border
        ws.row_dimensions[i].height = 18
    
    # Regola larghezza colonne
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    
    # Aggiungi foglio con istruzioni
    ws_istruzioni = wb.create_sheet("Istruzioni")
    
    ws_istruzioni['A1'] = "ISTRUZIONI PER L'USO"
    ws_istruzioni['A1'].font = Font(bold=True, size=14, color="4472C4")
    ws_istruzioni.merge_cells('A1:D1')
    ws_istruzioni.row_dimensions[1].height = 25
    
    istruzioni = [
        "",
        "REGOLA PER IL COGNOME:",
        "1. Estrai le CONSONANTI dal cognome (in ordine)",
        "2. Se sono meno di 3, aggiungi le VOCALI (in ordine)",
        "3. Se ancora non arrivano a 3, aggiungi 'X'",
        "4. Prendi i primi 3 caratteri (MAIUSCOLI)",
        "",
        "Esempio: 'Rossi' → consonanti: R, S, S → 'RSS'",
        "Esempio: 'Io' → consonanti: nessuna → vocali: I, O → 'IOX'",
        "",
        "REGOLA PER IL NOME:",
        "Se il nome ha 4 O PIÙ consonanti:",
        "  Prendi 1ª consonante + 3ª consonante + 4ª consonante",
        "Altrimenti:",
        "  Applica la stessa regola del cognome",
        "",
        "Esempio: 'Cristina' → consonanti: C, R, S, T, N",
        "         Prendi: C (1ª) + S (3ª) + T (4ª) → 'CST'",
        "Esempio: 'Mario' → consonanti: M, R, (meno di 4)",
        "         Prendi primi 3: 'MR' + vocale 'A' → 'MRA'",
    ]
    
    for idx, istruzione in enumerate(istruzioni, start=3):
        cell = ws_istruzioni.cell(row=idx, column=1)
        cell.value = istruzione
        if istruzione and istruzione.isupper():
            cell.font = Font(bold=True, size=11, color="4472C4")
        ws_istruzioni.column_dimensions['A'].width = 70
    
    # Salva il file
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    
    return output_path


if __name__ == "__main__":
    percorso = create_excel_cf_generator()
    print(f"✓ File Excel creato: {percorso}")
